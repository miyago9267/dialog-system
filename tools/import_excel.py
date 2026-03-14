"""從 Excel 劇情檔匯入，產出 zh_tw.json 翻譯條目、start.mcfunction 時間軸檔案、static txt 備份。

動作和站位會被解析為時間軸軌道和場景設置指令，而非純註解。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

# ── 角色顏色表（與 gen_localization.py 同步） ──────────────────────

COLOR_MAP: dict[str, str] = {
    "米迦爾": "§7【米迦爾】§r:",
    "尤尼恩": "§2【尤尼恩】§r:",
    "主角": "§b【%s§b】§r:",
    "約古": "§7【約古】§r:",
    "??": "§7【???】§r:",
    "???": "§7【???】§r:",
    "????": "§7【???】§r:",
    "?????": "§7【???】§r:",
    "賽勒尼亞": "§4【賽勒尼亞】§r:",
    "自律水晶防禦裝置": "§8【自律水晶防禦裝置】§r:",
    "菲恩特": "§3【菲恩特】§r:",
    "奈迪拉提雅": "§d【奈迪拉提雅】§r:",
    "史利亞卡": "§4【史利亞卡】§r:",
    "女性的聲音": "§d【女性的聲音】§r:",
    "年幼的男孩": "§e【年幼的男孩】§r:",
    "人類": "§7【人類】§r:",
    "堯稚克": "§4【堯稚克】§r:",
    "眾人": "§7【眾人】§r:",
    "荒蛛群": "§7【荒蛛群】§r:",
    "法羅涅拉": "§3【法羅涅拉】§r:",
    "阿多賽忒喀": "§c【阿多賽忒喀】§r:",
    "阿多賽忒克": "§c【阿多賽忒克】§r:",
    "「奧婕提亞」": "§5【「奧婕提亞」】§r:",
    "特提卡": "§7【特提卡】§r:",
    "阿帝嘉": "§7【阿帝嘉】§r:",
    "阿帝嘉小聲說道": "§7【阿帝嘉】§r:",
    "鐵匠賽克": "§7【鐵匠賽克】§r:",
    "賽克": "§7【鐵匠賽克】§r:",
    "麥洛倪雅莎": "§6【麥洛倪雅莎】§r:",
    "賽芬妲斯特": "§e【賽芬妲斯特】§r:",
    "克德": "§7【克德】§r:",
    "男子的聲音": "§7【男子的聲音】§r:",
    "年輕男性的聲音": "§7【年輕男性的聲音】§r:",
    "年邁男子的聲音": "§7【年邁男子的聲音】§r:",
    "巴柏諾斯": "§7【巴柏諾斯】§r:",
    "泰爾森": "§7【泰爾森】§r:",
    "丹德邦多": "§7【丹德邦多】§r:",
    "埃索拉": "§7【埃索拉】§r:",
    "女孩": "§d【女孩】§r:",
    "艾弗傑特": "§5【艾弗傑特】§r:",
    "莎琳絲": "§5【莎琳絲】§r:",
    "莎琳瑟芬": "§5【莎琳瑟芬】§r:",
    "希培利爾": "§7【希培利爾】§r:",
    "別瓦拉": "§7【別瓦拉】§r:",
    "迷霧使者": "§7【迷霧使者】§r:",
    "倪安": "§d【倪安】§r:",
    "守衛": "§7【守衛】§r:",
    "麥莎": "§6【麥洛倪雅莎】§r:",
}

# 認定為「主角台詞」的說話者
PLAYER_SPEAKERS = {"主角"}

# 偵測文本欄位的優先順序
TEXT_COL_CANDIDATES = ["新版劇情", "區域流程", "劇情內容"]

# 非劇情分頁（跳過）
SKIP_SHEETS = {"角色設定", "憶念存錄", "概述與音樂", "紀事", "範本", "支線劇情", "維蘭托恩尼", "多利亞爾", "帕蘭帝亞"}

# 分頁名稱 → chapter 代碼（用於場景 ID 無法推導 chapter 時的 fallback）
SHEET_CHAPTER_MAP: dict[str, str] = {
    "阿克丹佛": "village",
    "阿克丹佛(堯稚克)": "village_boss",
    "安特利亞斯": "sky_palace",
    "火源裂洞": "fire",
    "水源大堂": "water",
    "木源林間": "grass",
    "光源迷樓": "light",
    "泰洛科": "ghost_town",
    "尼哀爾": "soul_gravel",
    "坦塞安": "wild_valley",
    "柴哈卓": "lost_road",
    "朗拉多恩提": "sky_lands",
    "彌賽邇": "tower",
    "艾弗傑特": "forget",
    "伊特諾": "palace",
    "埃薩勒": "sunset_realm",
    "堯稚克": "dead_kingdom",
    "薩曼洛斯特": "samanrost",
    "維蘭托恩尼": "vilantonni",
    "多利亞爾": "doliar",
    "帕蘭帝亞": "palantia",
}

# 非合法場景 ID（跳過）
SKIP_SCENE_IDS = {"物品筆記", "999", "天-X"}

# 場景 ID 前綴正規化（Excel 名稱 → 程式碼名稱）
ID_PREFIX_NORMALIZE: dict[str, str] = {
    "skylands": "sky_lands",
    "soul_grave": "soul_gravel",
}

# 站位中視為「主角」的關鍵字
PLAYER_STAGING_NAMES = {"主角", "主角視角"}


# ── 資料結構 ────────────────────────────────────────────────────

@dataclass
class StagingEntry:
    """解析後的單個角色站位。"""
    name: str
    x: int | float | None = None
    y: int | float | None = None
    z: int | float | None = None
    facing: str = ""  # 面向目標或座標
    raw: str = ""     # 原始文字


@dataclass
class ActionEntry:
    """與特定台詞行關聯的動作。"""
    line_idx: int     # 對應的台詞行索引（0-based）
    text: str         # 動作描述
    is_cleanup: bool = False  # 是否為收場動作（畫面轉暗等）


# ── 工具函式 ────────────────────────────────────────────────────

def find_col_index(
    headers: list[str | None],
    candidates: list[str],
    rows: list[tuple] | None = None,
) -> int | None:
    """按優先順序在 headers 中找到第一個匹配且有實際內容的欄位索引。"""
    for candidate in candidates:
        for i, h in enumerate(headers):
            if h and candidate in h:
                if rows is not None:
                    has_content = any(
                        i < len(r) and r[i] and str(r[i]).strip()
                        for r in rows[:50]
                    )
                    if not has_content:
                        continue
                return i
    return None


def find_col(headers: list[str | None], name: str) -> int | None:
    """找到包含 name 的欄位索引。"""
    for i, h in enumerate(headers):
        if h and name in h:
            return i
    return None


def sanitize_scene_id(raw_id: str) -> str:
    """清理場景 ID：移除換行、空白，並正規化已知前綴。"""
    clean = raw_id.replace("\n", "").replace("\r", "").replace(" ", "").strip()
    for old_prefix, new_prefix in ID_PREFIX_NORMALIZE.items():
        if clean.startswith(old_prefix):
            clean = new_prefix + clean[len(old_prefix):]
            break
    return clean


def scene_id_to_chapter(scene_id: str, sheet_chapter: str | None = None) -> str:
    """從場景 ID 推導 chapter 名稱。"""
    if re.match(r"^M\d+$", scene_id) and sheet_chapter:
        return sheet_chapter
    if scene_id.isdigit() and sheet_chapter:
        return sheet_chapter
    if not re.match(r"^[a-z_]", scene_id) and sheet_chapter:
        return sheet_chapter
    m = re.match(r"^(.+?)(\d+)$", scene_id)
    if m:
        return m.group(1).rstrip("_")
    return sheet_chapter or scene_id


def convert_line(raw: str) -> tuple[str, bool]:
    """解析一行台詞，回傳 (轉換後文字, 是否為主角台詞)。"""
    raw = raw.strip()
    if not raw:
        return "", False
    if raw[0] in "(*（":
        return raw, False
    speaker, sep, content = raw.partition(":")
    if not sep:
        speaker, sep, content = raw.partition("：")
    if not sep:
        return raw, False
    speaker = speaker.strip()
    content = content.strip()
    content = content.replace('"主角"', "%s").replace("\"主角\"", "%s")
    prefix = COLOR_MAP.get(speaker, f"§7【{speaker}】§r:")
    result = f"{prefix}{content}"
    # 含 %s 的台詞都需要 text_player（替換玩家名），不只主角說的
    needs_player = speaker in PLAYER_SPEAKERS or "%s" in result
    return result, needs_player


# ── 站位解析 ────────────────────────────────────────────────────

_STAGING_RE = re.compile(
    r"^(.+?)\s*(-?\d+)\s+(\d+)\s+(-?\d+)\s*(.*?)$"
)


def parse_staging(text: str) -> list[StagingEntry]:
    """解析站位文字，提取角色座標。

    格式範例：
      奈迪拉提雅-1749 74 1498 面向主角
      尤尼恩-1751 74 1502 面向奈迪拉提雅
      主角-1749 74 1504 面向奈迪拉提雅
    """
    entries: list[StagingEntry] = []
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = _STAGING_RE.match(line)
        if m:
            name = m.group(1).strip()
            facing_raw = m.group(5).strip()
            # 清理 facing：「看向」「面向」
            facing = facing_raw.replace("看向", "").replace("面向", "").strip()
            entries.append(StagingEntry(
                name=name,
                x=int(m.group(2)),
                y=int(m.group(3)),
                z=int(m.group(4)),
                facing=facing,
                raw=line,
            ))
        else:
            # 無法解析的行 → 保留原始文字
            entries.append(StagingEntry(name="", raw=line))
    return entries


def gen_staging_commands(entries: list[StagingEntry]) -> list[str]:
    """從站位資料產出 mcfunction 指令（可直接使用或作為模板）。"""
    lines: list[str] = []
    player_entry: StagingEntry | None = None
    npc_entries: list[StagingEntry] = []

    for e in entries:
        if not e.x:
            # 無法解析的備註
            if e.raw:
                lines.append(f"# {e.raw}")
            continue
        if e.name in PLAYER_STAGING_NAMES:
            player_entry = e
        else:
            npc_entries.append(e)

    # 鎖定玩家（常見模式）
    lines.append("effect give @a slowness 999999 255 true")
    lines.append("effect give @a jump_boost 999999 128 true")
    lines.append("effect give @a blindness 1 0 true")
    lines.append("")

    # NPC 站位
    for e in npc_entries:
        facing_comment = f" (面向{e.facing})" if e.facing else ""
        lines.append(f"# TODO: summon {e.name} at {e.x} {e.y} {e.z}{facing_comment}")

    # 玩家傳送
    if player_entry:
        facing_target = ""
        if player_entry.facing:
            # 嘗試找到 facing 目標的座標
            for e in npc_entries:
                if player_entry.facing in e.name:
                    facing_target = f" facing {e.x} {e.y} {e.z}"
                    break
        if not facing_target and npc_entries:
            # 預設面向第一個 NPC
            e0 = npc_entries[0]
            facing_target = f" facing {e0.x} {e0.y} {e0.z}"
        lines.append(f"tp @a {player_entry.x} {player_entry.y} {player_entry.z}{facing_target}")

    return lines


# ── 動作解析 ────────────────────────────────────────────────────

_CLEANUP_KEYWORDS = ["畫面轉暗", "角色消失", "轉場"]
_SKIP_ACTIONS = ["修改"]


def is_cleanup_action(text: str) -> bool:
    """判斷動作是否為收場動作。"""
    return any(kw in text for kw in _CLEANUP_KEYWORDS)


def should_skip_action(text: str) -> bool:
    """判斷是否該跳過此動作。"""
    return text.strip() in _SKIP_ACTIONS


# ── 場景類 ──────────────────────────────────────────────────────

class Scene:
    def __init__(self, scene_id: str, chapter: str):
        self.scene_id = scene_id
        self.chapter = chapter
        self.lines: list[tuple[str, bool]] = []      # (translated_text, is_player)
        self.raw_lines: list[str] = []                # 原始台詞文本
        self.staging_entries: list[StagingEntry] = []  # 站位資料
        self.staging_raw: list[str] = []               # 站位原始文字
        self.actions: dict[int, list[str]] = {}        # line_idx → [action_texts]

    def set_staging(self, raw: str):
        """設定站位資訊（只在場景首行設定一次）。"""
        if raw and not self.staging_raw:
            self.staging_raw.append(raw.strip())
            self.staging_entries = parse_staging(raw)

    def add_line(self, raw: str, staging: str = "", action: str = ""):
        raw = raw.strip()
        if not raw:
            return
        translated, is_player = convert_line(raw)
        if not translated:
            return

        line_idx = len(self.lines)
        self.lines.append((translated, is_player))
        self.raw_lines.append(raw)

        if staging:
            self.set_staging(staging)

        if action:
            action = action.strip()
            if not should_skip_action(action):
                self.actions.setdefault(line_idx, []).append(action)

    def get_action_events(self) -> list[ActionEntry]:
        """取得所有動作事件，按台詞行索引排序。"""
        events: list[ActionEntry] = []
        for idx in sorted(self.actions.keys()):
            for text in self.actions[idx]:
                events.append(ActionEntry(
                    line_idx=idx,
                    text=text,
                    is_cleanup=is_cleanup_action(text),
                ))
        return events


# ── 分頁解析 ────────────────────────────────────────────────────

def parse_sheet(ws, sheet_name: str) -> list[Scene]:
    """解析一個分頁，回傳場景列表。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(c).strip() if c else None for c in rows[0]]

    scene_col = find_col(headers, "劇情編號")
    if scene_col is None:
        print(f"[SKIP] {sheet_name}: 無「劇情編號」欄", file=sys.stderr)
        return []

    data_rows = rows[1:]
    text_col = find_col_index(headers, TEXT_COL_CANDIDATES, rows=data_rows)
    if text_col is None:
        print(f"[SKIP] {sheet_name}: 找不到文本欄位", file=sys.stderr)
        return []

    staging_col = find_col(headers, "站位")
    action_col = find_col(headers, "動作")

    detected_col_name = headers[text_col] if text_col < len(headers) else ""
    print(f"[INFO] {sheet_name}: 使用欄位「{detected_col_name}」(col {text_col})", file=sys.stderr)

    sheet_chapter = SHEET_CHAPTER_MAP.get(sheet_name)
    scenes: list[Scene] = []
    current: Scene | None = None

    for row in data_rows:
        raw_id = str(row[scene_col]).strip() if scene_col < len(row) and row[scene_col] else ""
        cell_text = str(row[text_col]).strip() if text_col < len(row) and row[text_col] else ""
        cell_staging = ""
        cell_action = ""
        if staging_col is not None and staging_col < len(row) and row[staging_col]:
            cell_staging = str(row[staging_col])
        if action_col is not None and action_col < len(row) and row[action_col]:
            cell_action = str(row[action_col])

        # 新場景開始
        if raw_id and raw_id.lower() != "none":
            scene_id = sanitize_scene_id(raw_id)
            if scene_id in SKIP_SCENE_IDS:
                current = None
                continue
            chapter = scene_id_to_chapter(scene_id, sheet_chapter)
            current = Scene(scene_id, chapter)
            scenes.append(current)

        if current and cell_text:
            for sub_line in cell_text.split("\n"):
                sub_line = sub_line.strip()
                if sub_line:
                    current.add_line(sub_line, cell_staging, cell_action)
                    cell_staging = ""
                    cell_action = ""

    return scenes


# ── 輸出產生 ────────────────────────────────────────────────────

def gen_json_entries(scene: Scene) -> dict[str, str]:
    """產生 zh_tw.json 翻譯條目。"""
    entries = {}
    for idx, (text, _) in enumerate(scene.lines, start=1):
        key = f"story.{scene.chapter}.{scene.scene_id}.line{idx}"
        entries[key] = text
    return entries


def gen_mcfunction(scene: Scene, interval: int = 40) -> str:
    """產生 start.mcfunction 內容，含站位設置和動作軌道。"""
    out: list[str] = []

    # 觸發標記
    out.append("# 設置已觸發，防止重複執行")
    out.append(f"scoreboard players set {scene.scene_id}_triggered {scene.chapter}_story 1")
    out.append("")

    # ── 站位設置 ──
    if scene.staging_entries:
        out.append("# ── 場景設置（站位） ──────────────────────────────────────────")
        staging_cmds = gen_staging_commands(scene.staging_entries)
        out.extend(staging_cmds)
        out.append("")

    # ── 時間軸資料 ──
    if not scene.lines:
        return "\n".join(out)

    out.append("# ── 時間軸資料 ──────────────────────────────────────────────")

    # text 軌
    out.append(f"# text 軌（每行 {interval} ticks，*_player 表示帶玩家名稱）")
    text_events = []
    for idx, (_, is_player) in enumerate(scene.lines):
        t = idx * interval
        typ = "text_player" if is_player else "text"
        key = f"story.{scene.chapter}.{scene.scene_id}.line{idx + 1}"
        text_events.append('{' + f't:{t},type:"{typ}",key:"{key}"' + '}')
    out.append(f"data modify storage dialogtest:story run.text set value [{','.join(text_events)}]")
    out.append("")

    # action 軌
    action_events = scene.get_action_events()
    non_cleanup = [a for a in action_events if not a.is_cleanup]
    cleanup_actions = [a for a in action_events if a.is_cleanup]

    if non_cleanup:
        out.append(f"# [DISABLED] action 軌（動作事件，每個指向 stub mcfunction）")
        act_items = []
        for i, act in enumerate(non_cleanup, 1):
            t = act.line_idx * interval
            fn = f"dialogtest:{scene.chapter}/{scene.scene_id}/act{i}"
            act_items.append('{' + f't:{t},type:"fn",fn:"{fn}"' + '}')
        out.append(f"# data modify storage dialogtest:story run.action set value [{','.join(act_items)}]")
        # 列出動作對照表
        for i, act in enumerate(non_cleanup, 1):
            t = act.line_idx * interval
            raw_line = scene.raw_lines[act.line_idx] if act.line_idx < len(scene.raw_lines) else ""
            out.append(f"# act{i} (t={t}, line{act.line_idx+1}): {act.text}")
        out.append("")

    # ctrl 軌
    end_t = len(scene.lines) * interval
    cleanup_fn = f"dialogtest:{scene.chapter}/{scene.scene_id}/cleanup" if cleanup_actions else "dialogtest:operations/timeline/end"
    out.append(f"# ctrl 軌：最後一行後 {interval} ticks 結束場景")
    out.append(
        f'data modify storage dialogtest:story run.ctrl set value '
        f'[{{t:{end_t},type:"fn",fn:"{cleanup_fn}"}}]'
    )
    out.append("")

    # ── 啟動時間軸 ──
    out.append("# ── 啟動時間軸 ──────────────────────────────────────────────")
    out.append("scoreboard players set _scene_tick dialog_timer 0")
    out.append("data modify storage dialogtest:story run.playing set value 1b")
    out.append('data modify storage dialogtest:story run.mode set value "timeline"')
    out.append(f'data modify storage dialogtest:story run.scene set value "{scene.chapter}_{scene.scene_id}"')
    out.append("data modify storage dialogtest:story run.scene_tick set value 0")
    out.append("")

    return "\n".join(out)


def gen_action_stubs(scene: Scene, interval: int = 40) -> dict[str, str]:
    """產生動作 stub mcfunction 檔案。回傳 {相對路徑: 內容}。"""
    stubs: dict[str, str] = {}
    action_events = scene.get_action_events()
    non_cleanup = [a for a in action_events if not a.is_cleanup]
    cleanup_actions = [a for a in action_events if a.is_cleanup]

    for i, act in enumerate(non_cleanup, 1):
        t = act.line_idx * interval
        raw_line = scene.raw_lines[act.line_idx] if act.line_idx < len(scene.raw_lines) else ""
        path = f"{scene.chapter}/{scene.scene_id}/act{i}.mcfunction"
        content = (
            f"# 動作：{act.text}\n"
            f"# 觸發時機：t={t} (line{act.line_idx + 1})\n"
            f"# 對應台詞：{raw_line}\n"
            f"# TODO: 實作此動作\n"
        )
        stubs[path] = content

    if cleanup_actions:
        path = f"{scene.chapter}/{scene.scene_id}/cleanup.mcfunction"
        descriptions = "\n".join(f"# - {a.text}" for a in cleanup_actions)
        content = (
            f"# 收場動作\n"
            f"{descriptions}\n"
            f"\n"
            f"# 停止時間軸\n"
            f"function dialogtest:operations/timeline/end\n"
            f"\n"
            f"# TODO: 清除場景實體和效果\n"
            f"# effect clear @a slowness\n"
            f"# effect clear @a jump_boost\n"
            f"# execute as @e[tag={scene.scene_id}] run kill @s\n"
        )
        stubs[path] = content

    return stubs


def gen_static_txt(scene: Scene) -> str:
    """產生 static txt 純文字備份。"""
    return "\n".join(scene.raw_lines) + "\n" if scene.raw_lines else ""


# ── 主程式 ──────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="從 Excel 劇情檔匯入，產出 json + mcfunction + static txt。")
    parser.add_argument("--excel", required=True, help="Excel 檔案路徑")
    parser.add_argument("--output-json", default="tools/output.json", help="翻譯 JSON 輸出路徑")
    parser.add_argument("--output-functions", default="data/dialogtest/functions", help="mcfunction 輸出根目錄")
    parser.add_argument("--output-static", default="data/static", help="static txt 輸出根目錄")
    parser.add_argument("--sheets", default="", help="只處理指定分頁（逗號分隔），留空處理全部")
    parser.add_argument("--dry-run", action="store_true", help="預覽模式，不寫入檔案")
    parser.add_argument("--no-overwrite", action="store_true", help="不覆蓋已有的 start.mcfunction")
    parser.add_argument("--interval", type=int, default=40, help="台詞間隔 ticks（預設 40）")
    args = parser.parse_args()

    excel_path = Path(args.excel).resolve()
    if not excel_path.exists():
        raise SystemExit(f"Excel 檔案不存在: {excel_path}")

    base_dir = excel_path.parent
    out_json = (base_dir / args.output_json).resolve()
    out_fn = (base_dir / args.output_functions).resolve()
    out_static = (base_dir / args.output_static).resolve()

    filter_sheets = set(s.strip() for s in args.sheets.split(",") if s.strip()) if args.sheets else None

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    all_json: dict[str, str] = {}
    total_scenes = 0
    total_lines = 0
    total_stubs = 0

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        if filter_sheets and sheet_name not in filter_sheets:
            continue

        ws = wb[sheet_name]
        scenes = parse_sheet(ws, sheet_name)

        for scene in scenes:
            if not scene.lines:
                print(f"[WARN] {sheet_name}/{scene.scene_id}: 無台詞，跳過", file=sys.stderr)
                continue

            total_scenes += 1
            total_lines += len(scene.lines)

            # JSON
            entries = gen_json_entries(scene)
            all_json.update(entries)

            # mcfunction
            mcfn = gen_mcfunction(scene, interval=args.interval)
            mcfn_path = out_fn / scene.chapter / scene.scene_id / "start.mcfunction"

            # action stubs
            stubs = gen_action_stubs(scene, interval=args.interval)
            total_stubs += len(stubs)

            # static txt
            txt = gen_static_txt(scene)
            txt_path = out_static / scene.chapter / f"{scene.scene_id}.txt"

            if args.dry_run:
                print(f"\n{'='*60}")
                print(f"[DRY-RUN] {sheet_name} → {scene.chapter}/{scene.scene_id} ({len(scene.lines)} lines, {len(stubs)} stubs)")
                print(f"  mcfunction: {mcfn_path}")
                if stubs:
                    for sp in stubs:
                        print(f"  stub: {sp}")
                # 顯示前 3 行 JSON
                for i, (k, v) in enumerate(entries.items()):
                    if i >= 3:
                        print(f"  ... ({len(entries) - 3} more)")
                        break
                    print(f"  {k}: {v[:60]}{'...' if len(v) > 60 else ''}")
                # 顯示動作對照
                action_events = scene.get_action_events()
                for act in action_events:
                    t = act.line_idx * args.interval
                    tag = "[cleanup]" if act.is_cleanup else ""
                    print(f"  action t={t} line{act.line_idx+1}{tag}: {act.text}")
            else:
                # 寫入 mcfunction
                if args.no_overwrite and mcfn_path.exists():
                    print(f"[SKIP] {mcfn_path} 已存在，跳過", file=sys.stderr)
                else:
                    mcfn_path.parent.mkdir(parents=True, exist_ok=True)
                    mcfn_path.write_text(mcfn, encoding="utf-8")

                # 寫入 stubs（不覆蓋已有檔案）
                for stub_rel, stub_content in stubs.items():
                    stub_path = out_fn / stub_rel
                    if not stub_path.exists():
                        stub_path.parent.mkdir(parents=True, exist_ok=True)
                        stub_path.write_text(stub_content, encoding="utf-8")

                # 寫入 static txt
                txt_path.parent.mkdir(parents=True, exist_ok=True)
                txt_path.write_text(txt, encoding="utf-8")

    wb.close()

    # 寫入 JSON
    if all_json:
        if args.dry_run:
            print(f"\n{'='*60}")
            print(f"[DRY-RUN] JSON: {len(all_json)} entries → {out_json}")
        else:
            out_json.parent.mkdir(parents=True, exist_ok=True)
            out_json.write_text(
                json.dumps(all_json, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            print(f"[OK] JSON: {len(all_json)} entries → {out_json}")

    print(
        f"\n[DONE] {total_scenes} scenes, {total_lines} lines, {total_stubs} action stubs processed.",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
